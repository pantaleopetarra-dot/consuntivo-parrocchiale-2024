from openpyxl import Workbook
from io import BytesIO
from .models import BilancioConsuntivo, Entrata, Uscita

def genera_excel_consuntivo(bilancio):
    """
    Genera un file Excel identico al modello fornito,
    con fogli ENTRATE, USCITE e GESTIONI SPECIALI.
    """
    wb = Workbook()

    # Foglio ENTRATE
    ws_ent = wb.active
    ws_ent.title = "ENTRATE"
    ws_ent.append(["Giorno", "Mese", "Codice*", "Causale 1", "Causale 2", "Causale 3", "Importo Euro", "", "* Inserire i codici del piano dei conti"])
    for entrata in bilancio.entrate.all():
        ws_ent.append([
            entrata.giorno or "",
            entrata.mese,
            entrata.codice,
            entrata.causale1,
            entrata.causale2,
            entrata.causale3,
            float(entrata.importo)
        ])
    ws_ent.append(["", "", "", "", "", "TOTALE ENTRATE ANNO - 2024", f"€ {bilancio.totale_entrate:.2f}"])

    # Foglio USCITE
    ws_usc = wb.create_sheet("USCITE")
    ws_usc.append(["Giorno", "Mese", "Codice*", "Causale 1", "Causale 2", "Causale 3", "Importo Euro", "", "* Inserire i codici del piano dei conti"])
    for uscita in bilancio.uscite.all():
        ws_usc.append([
            uscita.giorno or "",
            uscita.mese,
            uscita.codice,
            uscita.causale1,
            uscita.causale2,
            uscita.causale3,
            float(uscita.importo)
        ])
    ws_usc.append(["", "", "", "", "", "TOTALE USCITE ANNO - 2024", f"€ {bilancio.totale_uscite:.2f}"])

    # Foglio GESTIONI SPECIALI
    ws_gs = wb.create_sheet("GESTIONI SPECIALI")
    ws_gs.append(["Anno", "", "Totale Generale", "€ 0.00"])
    ws_gs.append(["", "", "Totale Parziale", "€ 0.00"])
    ws_gs.append(["Giorno", "Mese", "Codice*", "Causale 1", "Causale 2", "Causale 3", "Importo Euro", "", "* INSERIRE I CODICI LETTERALI DELLA PAGINA 5"])

    # Codici gestioni speciali (A, B, C, D, ecc.)
    codici_speciali = [
        ('A', 'per la cooperazione diocesana'),
        ('B', 'per la quaresima di fraternita'),
        ('C', 'per la terra santa'),
        ('D', 'per la carità del papa (obolo di s. pietro)'),
        ('E', 'per le missioni'),
        ('F', 'per il seminario'),
        ('G', 'per s. vincenzo'),
        ('H', 'per caritas'),
        ('I', 'carità parrocchiale'),
        ('L', 'altre'),
    ]

    totale_generale = 0
    for codice, descr in codici_speciali:
        totale = bilancio.entrate.filter(codice=codice).aggregate(total=models.Sum('importo'))['total'] or 0
        ws_gs.append(["", "", codice, descr, "", "", f"€ {totale:.2f}"])
        totale_generale += totale

    # Aggiorna i totali
    ws_gs.cell(row=1, column=4).value = f"€ {totale_generale:.2f}"
    ws_gs.cell(row=2, column=4).value = f"€ {totale_generale:.2f}"

    # Salva in memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer